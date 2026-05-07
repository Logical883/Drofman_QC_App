"""
CVAT QC Metrics Flask Application
"""
import os
import io
import uuid
import json
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from metrics import detect_and_parse, compute_metrics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024

ALLOWED = {".xml", ".json"}


def allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED


def serialize_report(report, annotator_name):
    frame_data = [
        {
            "frame_id": fr.frame_id,
            "tp": fr.tp,
            "fp": fr.fp,
            "fn": fr.fn,
            "mean_iou": round(sum(fr.iou_scores) / len(fr.iou_scores) * 100, 1) if fr.iou_scores else 0,
        }
        for fr in report.frame_results
    ]
    return {
        "annotator": annotator_name,
        "precision": report.precision,
        "recall": report.recall,
        "accuracy": report.accuracy,
        "mean_iou": report.mean_iou,
        "tp": report.tp,
        "fp": report.fp,
        "fn": report.fn,
        "iou_threshold": report.iou_threshold,
        "total_gt": report.total_gt,
        "total_pred": report.total_pred,
        "matched_frames": report.matched_frames,
        "gt_only_frames": report.gt_only_frames,
        "pred_only_frames": report.pred_only_frames,
        "format_gt": report.format_gt,
        "format_pred": report.format_pred,
        "detected_name": annotator_name,
        "frame_results": frame_data,
        "label_breakdown": report.label_breakdown,
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    gt_file = request.files.get("gt_file")
    pred_file = request.files.get("pred_file")

    if not gt_file or not pred_file:
        return jsonify({"error": "Both ground truth and prediction files are required."}), 400
    if not allowed_file(gt_file.filename) or not allowed_file(pred_file.filename):
        return jsonify({"error": "Only .xml and .json files are supported."}), 400

    iou_threshold = float(request.form.get("iou_threshold", 50)) / 100
    compare_labels = request.form.get("compare_labels", "true") == "true"

    uid = uuid.uuid4().hex
    gt_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_gt_{secure_filename(gt_file.filename)}")
    pred_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_pred_{secure_filename(pred_file.filename)}")

    try:
        gt_file.save(gt_path)
        pred_file.save(pred_path)
        gt_frames, gt_fmt, _ = detect_and_parse(gt_path)
        pred_frames, pred_fmt, detected_name = detect_and_parse(pred_path)
        report = compute_metrics(gt_frames, pred_frames,
                                 iou_threshold=iou_threshold,
                                 compare_labels=compare_labels)
        report.format_gt = gt_fmt
        report.format_pred = pred_fmt
        return jsonify(serialize_report(report, detected_name or "Annotator"))
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        for p in [gt_path, pred_path]:
            if os.path.exists(p):
                os.remove(p)


@app.route("/analyze-batch", methods=["POST"])
def analyze_batch():
    gt_file = request.files.get("gt_file")
    if not gt_file:
        return jsonify({"error": "Ground truth file is required."}), 400

    iou_threshold = float(request.form.get("iou_threshold", 50)) / 100
    compare_labels = request.form.get("compare_labels", "true") == "true"
    pred_files = request.files.getlist("pred_files")
    annotator_names = request.form.getlist("annotator_names")

    if not pred_files:
        return jsonify({"error": "At least one annotator file is required."}), 400

    uid = uuid.uuid4().hex
    gt_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_gt_{secure_filename(gt_file.filename)}")

    try:
        gt_file.save(gt_path)
        gt_frames, gt_fmt, _ = detect_and_parse(gt_path)
        results = []
        errors = []

        for i, pf in enumerate(pred_files):
            name = annotator_names[i] if i < len(annotator_names) and annotator_names[i].strip() else f"Annotator {i+1}"
            if not allowed_file(pf.filename):
                errors.append(f"{name}: unsupported file format")
                continue
            pred_path = os.path.join(app.config["UPLOAD_FOLDER"],
                                     f"{uid}_pred{i}_{secure_filename(pf.filename)}")
            try:
                pf.save(pred_path)
                pred_frames, pred_fmt, detected_name = detect_and_parse(pred_path)
                # Auto-fill name from file if user left the field blank
                if not annotator_names[i].strip() and detected_name:
                    name = detected_name
                report = compute_metrics(gt_frames, pred_frames,
                                         iou_threshold=iou_threshold,
                                         compare_labels=compare_labels)
                report.format_gt = gt_fmt
                report.format_pred = pred_fmt
                results.append(serialize_report(report, name))
            except Exception as e:
                errors.append(f"{name}: {str(e)}")
            finally:
                if os.path.exists(pred_path):
                    os.remove(pred_path)

        return jsonify({"results": results, "errors": errors})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if os.path.exists(gt_path):
            os.remove(gt_path)


@app.route("/export-excel", methods=["POST"])
def export_excel():
    data = request.get_json()
    results = data.get("results", [])
    iou_threshold = data.get("iou_threshold", 50)
    session_name = data.get("session_name", "QC Report")

    if not results:
        return jsonify({"error": "No results to export."}), 400

    wb = Workbook()

    C_HEADER_BG   = "1A2332"
    C_HEADER_FG   = "FFFFFF"
    C_ACCENT      = "00C8A0"
    C_ACCENT_DARK = "008F72"
    C_GOOD_BG     = "D6F5ED"
    C_GOOD_FG     = "0A6B4A"
    C_WARN_BG     = "FFF3CD"
    C_WARN_FG     = "7D5A00"
    C_BAD_BG      = "FFE0E0"
    C_BAD_FG      = "8B1A1A"
    C_ROW_ALT     = "F7FFFE"
    C_BORDER      = "CCCCCC"

    def hdr_font(size=11, bold=True, color=C_HEADER_FG):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border_all():
        return Border(
            left=Side(style="thin", color=C_BORDER),
            right=Side(style="thin", color=C_BORDER),
            top=Side(style="thin", color=C_BORDER),
            bottom=Side(style="thin", color=C_BORDER))

    def score_style(val):
        if val >= 80:   return C_GOOD_BG, C_GOOD_FG
        elif val >= 50: return C_WARN_BG, C_WARN_FG
        return C_BAD_BG, C_BAD_FG

    def set_cell(ws, row, col, value, font=None, bg=None, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        if font: c.font = font
        if bg:   c.fill = fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border_all()
        return c

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[5].height = 32

    ws.merge_cells("A1:L1")
    ws["A1"].value = f"CVAT Quality Control — {session_name}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].border = Border(bottom=Side(style="medium", color=C_ACCENT))

    ws.merge_cells("A2:L2")
    ws["A2"].value = (f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}   |   "
                      f"IoU Threshold: {iou_threshold}%   |   Annotators: {len(results)}   |   "
                      f"Smart frame matching: ON (GT-only frames excluded per annotator)")
    ws["A2"].font = Font(name="Arial", size=9, color="888888")
    ws["A2"].fill = fill(C_HEADER_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    for r in [3, 4]:
        ws.merge_cells(f"A{r}:L{r}")
        ws[f"A{r}"].fill = fill(C_HEADER_BG if r == 3 else "F0F0F0")

    headers = ["Rank", "Annotator Name", "Precision (%)", "Recall (%)",
               "Accuracy/F1 (%)", "Mean IoU (%)", "TP", "FP", "FN",
               "Scored Frames", "GT-Only Frames Skipped", "Total GT Boxes"]
    col_widths = [8, 24, 16, 14, 17, 14, 8, 8, 8, 15, 24, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = hdr_font(10)
        c.fill = fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=C_ACCENT),
                          left=Side(style="thin", color="444444"),
                          right=Side(style="thin", color="444444"))
        ws.column_dimensions[get_column_letter(col)].width = w

    sorted_results = sorted(results, key=lambda r: r["accuracy"], reverse=True)

    for i, r in enumerate(sorted_results):
        row = 6 + i
        ws.row_dimensions[row].height = 22
        row_bg = "FFFFFF" if i % 2 == 0 else C_ROW_ALT
        medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(i + 1, str(i + 1))

        values = [medal, r["annotator"], r["precision"], r["recall"],
                  r["accuracy"], r["mean_iou"], r["tp"], r["fp"], r["fn"],
                  r.get("matched_frames", "—"),
                  r.get("gt_only_frames", "—"),
                  r["total_gt"]]
        aligns = ["center", "left", "center", "center", "center", "center",
                  "center", "center", "center", "center", "center", "center"]

        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = cell_font(10, bold=(col == 2))
            c.alignment = Alignment(horizontal=aln, vertical="center")
            c.border = border_all()
            if col in (3, 4, 5, 6) and isinstance(val, (int, float)):
                bg, fg = score_style(val)
                c.fill = fill(bg)
                c.font = Font(name="Arial", size=10, bold=True, color=fg)
                c.number_format = '0.00"%"'
            else:
                c.fill = fill(row_bg)

    # Team average row
    if len(results) > 1:
        avg_row = 6 + len(results)
        ws.row_dimensions[avg_row].height = 24
        avg = {k: sum(r[k] for r in results) / len(results)
               for k in ["precision", "recall", "accuracy", "mean_iou"]}
        avg_data = ["", "TEAM AVERAGE", avg["precision"], avg["recall"],
                    avg["accuracy"], avg["mean_iou"], "", "", "", "", "", ""]
        for col, val in enumerate(avg_data, 1):
            c = ws.cell(row=avg_row, column=col, value=val)
            c.font = Font(name="Arial", size=10, bold=True, color=C_HEADER_FG)
            c.fill = fill(C_ACCENT_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(top=Side(style="medium", color="FFFFFF"))
            if col in (3, 4, 5, 6) and isinstance(val, float):
                c.number_format = '0.00"%"'

    ws.freeze_panes = "A6"

    # ── Per-annotator detail sheets ───────────────────────────────────────
    for r in results:
        safe_name = (r["annotator"][:28]
                     .replace("/", "-").replace("\\", "-").replace(":", "-")
                     .replace("*", "-").replace("?", "-")
                     .replace("[", "(").replace("]", ")"))
        ws2 = wb.create_sheet(title=safe_name)
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:H1")
        ws2["A1"].value = f"{r['annotator']} — Detailed Report"
        ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=C_HEADER_FG)
        ws2["A1"].fill = fill(C_HEADER_BG)
        ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 34

        # Frame coverage info row
        ws2.merge_cells("A2:H2")
        mf = r.get("matched_frames", "?")
        sk = r.get("gt_only_frames", "?")
        ws2["A2"].value = (f"Scored on {mf} matched frame(s)   |   "
                           f"{sk} GT-only frame(s) excluded (not in annotator's job)   |   "
                           f"IoU Threshold: {r['iou_threshold']}%")
        ws2["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
        ws2["A2"].fill = fill("F5F5F5")
        ws2["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[2].height = 18

        # Metric cards
        ws2.row_dimensions[4].height = 28
        ws2.row_dimensions[5].height = 22
        metric_cards = [("Precision", r["precision"], 1),
                        ("Recall", r["recall"], 3),
                        ("Accuracy (F1)", r["accuracy"], 5),
                        ("Mean IoU", r["mean_iou"], 7)]
        for label, val, col in metric_cards:
            bg, fg = score_style(val)
            for row_n in [3, 4, 5]:
                ws2.merge_cells(start_row=row_n, start_column=col,
                                end_row=row_n, end_column=col + 1)
            ws2.cell(row=3, column=col, value=label).font = Font(name="Arial", size=9, color="888888")
            ws2.cell(row=3, column=col).fill = fill("F5F5F5")
            ws2.cell(row=3, column=col).alignment = Alignment(horizontal="center")
            vc = ws2.cell(row=4, column=col, value=val)
            vc.font = Font(name="Arial", size=18, bold=True, color=fg)
            vc.fill = fill(bg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.number_format = '0.00"%"'
            pc = ws2.cell(row=5, column=col, value="%")
            pc.font = Font(name="Arial", size=9, color=fg)
            pc.fill = fill(bg)
            pc.alignment = Alignment(horizontal="center")

        for col in range(1, 13):
            ws2.column_dimensions[get_column_letter(col)].width = 14

        # Per-frame table
        fhr = 7
        ws2.row_dimensions[fhr].height = 26
        for col, h in enumerate(["Frame", "TP", "FP", "FN", "Mean IoU (%)"], 1):
            c = ws2.cell(row=fhr, column=col, value=h)
            c.font = hdr_font(10)
            c.fill = fill(C_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(bottom=Side(style="medium", color=C_ACCENT))
        ws2.column_dimensions["A"].width = 20
        for col_i in range(2, 6):
            ws2.column_dimensions[get_column_letter(col_i)].width = 14

        for fi, fr in enumerate(r.get("frame_results", [])):
            frow = fhr + 1 + fi
            ws2.row_dimensions[frow].height = 20
            row_bg = "FFFFFF" if fi % 2 == 0 else C_ROW_ALT
            for col, val in enumerate([fr["frame_id"], fr["tp"], fr["fp"],
                                        fr["fn"], fr["mean_iou"]], 1):
                c = ws2.cell(row=frow, column=col, value=val)
                c.font = cell_font(9)
                c.fill = fill(row_bg)
                c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                        vertical="center")
                c.border = border_all()
                if col == 5:
                    c.number_format = '0.0"%"'

        # Per-label table
        if r.get("label_breakdown"):
            lsc = 7
            for col, h in enumerate(["Class", "TP", "FP", "FN",
                                      "Precision (%)", "Recall (%)"], lsc):
                c = ws2.cell(row=fhr, column=col, value=h)
                c.font = hdr_font(10)
                c.fill = fill(C_HEADER_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(bottom=Side(style="medium", color=C_ACCENT))
            for col_i, w in enumerate([18, 10, 10, 10, 16, 14], lsc):
                ws2.column_dimensions[get_column_letter(col_i)].width = w

            for li, (lbl, s) in enumerate(r["label_breakdown"].items()):
                lrow = fhr + 1 + li
                ws2.row_dimensions[lrow].height = 20
                row_bg = "FFFFFF" if li % 2 == 0 else C_ROW_ALT
                for col, val in enumerate([lbl, s["tp"], s["fp"], s["fn"],
                                           s["precision"], s["recall"]], lsc):
                    c = ws2.cell(row=lrow, column=col, value=val)
                    c.font = cell_font(9)
                    c.fill = fill(row_bg)
                    c.alignment = Alignment(
                        horizontal="center" if col > lsc else "left",
                        vertical="center")
                    c.border = border_all()
                    if col >= lsc + 4:
                        bg2, fg2 = score_style(val)
                        c.fill = fill(bg2)
                        c.font = Font(name="Arial", size=9, bold=True, color=fg2)
                        c.number_format = '0.0"%"'

    # ── Raw Data sheet ────────────────────────────────────────────────────
    ws_raw = wb.create_sheet(title="Raw Data")
    ws_raw.sheet_view.showGridLines = False
    raw_hdrs = ["Annotator", "Precision (%)", "Recall (%)", "Accuracy/F1 (%)",
                "Mean IoU (%)", "TP", "FP", "FN", "Scored Frames",
                "GT-Only Frames Skipped", "Total GT Boxes", "Total Pred Boxes",
                "IoU Threshold (%)"]
    for col, h in enumerate(raw_hdrs, 1):
        c = ws_raw.cell(row=1, column=col, value=h)
        c.font = hdr_font(10)
        c.fill = fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_raw.column_dimensions[get_column_letter(col)].width = 20

    for i, r in enumerate(results, 2):
        row_vals = [r["annotator"], r["precision"], r["recall"], r["accuracy"],
                    r["mean_iou"], r["tp"], r["fp"], r["fn"],
                    r.get("matched_frames", ""), r.get("gt_only_frames", ""),
                    r["total_gt"], r["total_pred"], r["iou_threshold"]]
        for col, val in enumerate(row_vals, 1):
            c = ws_raw.cell(row=i, column=col, value=val)
            c.font = cell_font(10)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                    vertical="center")
            c.border = border_all()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = (f"CVAT_QC_{session_name.replace(' ', '_')}_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


os.makedirs("uploads", exist_ok=True)

if __name__ == "__main__":
    app.run(debug=True, port=5050)